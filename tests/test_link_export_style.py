# ======================================
# Estilo de hipervínculos Link Nota / Link (Streaming)
# ======================================
import io
import unittest
from zipfile import ZipFile
from xml.etree import ElementTree as ET

from openpyxl import load_workbook

from pipeline import BASE_OUTPUT_COLUMNS, KEY_MAP, PLAIN_HYPERLINK_COLUMNS, generate_output_excel

NS_MAIN = {"m": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}


def _sample_rows():
    return [
        {
            "ID Noticia": 101,
            "Título": "Nota de prueba",
            "Link Nota": {"value": "Link", "url": "https://example.com/nota"},
            "Link (Streaming - Imagen)": {"value": "Link", "url": "https://example.com/stream"},
            "resumen corto": "texto",
        }
    ]


def _font_is_black(font) -> bool:
    color = getattr(font, "color", None)
    if color is None:
        return True
    rgb = getattr(color, "rgb", None)
    if rgb is None:
        theme = getattr(color, "theme", None)
        indexed = getattr(color, "indexed", None)
        if theme is None and indexed is None:
            return True
        return False
    hex_rgb = str(rgb).upper().replace(" ", "")
    return hex_rgb.endswith("000000")


def _font_not_underlined(font) -> bool:
    underline = getattr(font, "underline", None)
    return underline in (None, False, "none", "None")


def _cell_is_hyperlinked(cell) -> bool:
    value = str(cell.value or "")
    if "HYPERLINK(" in value.upper() and "http" in value.lower():
        return True
    if cell.hyperlink is not None:
        target = cell.hyperlink.target or cell.hyperlink.ref
        return str(target).startswith("http")
    return False


def _hyperlink_style_font(styles_xml: bytes):
    tree = ET.fromstring(styles_xml)
    styles = tree.findall("m:cellStyles/m:cellStyle", NS_MAIN)
    hyperlink = next((el for el in styles if el.attrib.get("name") == "Hyperlink"), None)
    if hyperlink is None:
        return None, None
    xf_id = int(hyperlink.attrib.get("xfId", "1"))
    style_xfs = tree.findall("m:cellStyleXfs/m:xf", NS_MAIN)
    font_id = int(style_xfs[xf_id].attrib.get("fontId", "0"))
    fonts = tree.findall("m:fonts/m:font", NS_MAIN)
    return hyperlink, fonts[font_id]


class LinkExportStyleTests(unittest.TestCase):
    def test_link_nota_and_streaming_are_black_without_underline(self):
        data = generate_output_excel(_sample_rows(), KEY_MAP)
        wb = load_workbook(io.BytesIO(data))
        ws = wb["Resultado"]
        headers = [cell.value for cell in ws[1]]

        for col_name in ("Link Nota", "Link (Streaming - Imagen)"):
            self.assertIn(col_name, PLAIN_HYPERLINK_COLUMNS)
            self.assertIn(col_name, headers)
            col_idx = headers.index(col_name) + 1
            cell = ws.cell(row=2, column=col_idx)
            self.assertTrue(_cell_is_hyperlinked(cell), f"{col_name} debe seguir siendo un hipervínculo")
            value = str(cell.value or "")
            self.assertIn("HYPERLINK(", value.upper())
            self.assertIn("https://example.com/", value)
            self.assertIn("Link", value)
            self.assertTrue(_font_is_black(cell.font), f"{col_name} debe ser texto negro, color={cell.font.color}")
            self.assertTrue(
                _font_not_underlined(cell.font),
                f"{col_name} no debe ir subrayado, underline={cell.font.underline}",
            )
            style_name = (cell.style or "") if isinstance(cell.style, str) else ""
            self.assertEqual(style_name, "Hyperlink")

        title_idx = headers.index("Título") + 1
        title_cell = ws.cell(row=2, column=title_idx)
        self.assertEqual(title_cell.value, "Nota de prueba")
        self.assertIsNone(title_cell.hyperlink)

    def test_xlsx_xml_overrides_hyperlink_style_and_uses_formula(self):
        data = generate_output_excel(_sample_rows(), KEY_MAP)
        with ZipFile(io.BytesIO(data)) as zf:
            sheet = zf.read("xl/worksheets/sheet1.xml")
            styles = zf.read("xl/styles.xml")

        sheet_tree = ET.fromstring(sheet)
        rel_hyperlinks = [
            el for el in sheet_tree.iter() if el.tag.endswith("hyperlink")
        ]
        self.assertEqual(
            len(rel_hyperlinks),
            0,
            "Link Nota / Streaming no deben usar <hyperlinks> de relación (Excel las restaura azules)",
        )

        formulas = [
            el.text or ""
            for el in sheet_tree.iter()
            if el.tag.endswith("}f") or el.tag == "f"
        ]
        joined = "\n".join(formulas)
        self.assertIn("HYPERLINK(", joined.upper())
        self.assertIn("https://example.com/nota", joined)
        self.assertIn("https://example.com/stream", joined)

        named, font = _hyperlink_style_font(styles)
        self.assertIsNotNone(named, "styles.xml debe definir el estilo nombrado Hyperlink")
        self.assertEqual(named.attrib.get("builtinId"), "8")
        color = font.find("m:color", NS_MAIN)
        underline = font.find("m:u", NS_MAIN)
        rgb = (color.attrib.get("rgb") if color is not None else "") or ""
        theme = color.attrib.get("theme") if color is not None else None
        self.assertTrue(rgb.upper().endswith("000000"), f"Hyperlink style debe ser negro, rgb={rgb} theme={theme}")
        self.assertIsNone(theme, "Hyperlink style no debe usar theme-10 azul")
        self.assertIsNone(underline, "Hyperlink style no debe llevar subrayado")

    def test_plain_hyperlink_columns_are_only_the_two_requested(self):
        self.assertEqual(
            PLAIN_HYPERLINK_COLUMNS,
            frozenset({"Link Nota", "Link (Streaming - Imagen)"}),
        )
        for col in PLAIN_HYPERLINK_COLUMNS:
            self.assertIn(col, BASE_OUTPUT_COLUMNS)


if __name__ == "__main__":
    unittest.main()
